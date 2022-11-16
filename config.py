import openpyxl
from urllib import request


url = "https://docs.google.com/spreadsheets/d/1LzTIXOesLNGy-9yDxF_2Xab9ESAoWdLj/" \
      "export?format=xlsx&id=1LzTIXOesLNGy-9yDxF_2Xab9ESAoWdLj"
request.urlretrieve(url, "Database/database.xlsx")


# BOT_TOKEN = '5687247228:AAFPTFmvJCtVU_weFVUnwCf0i7BPkMNAhiA'  # test_bot
BOT_TOKEN = '5221109021:AAGl2b6Vs9Id1Yzss89l-34uwMkQcFNnETQ'
BOT_TOKEN_logs = '5158112868:AAEKWw51sG5IT9Sxqbb4F1A6TTaaHOHyrQA'


excel_db = openpyxl.load_workbook('Database/database.xlsx')
ex_user = excel_db["Пользователи"]
ex_swats = excel_db["Спецназ"]
worksheet_build_pvp = excel_db["ПВП"]
worksheet_build_bg = excel_db["БГ"]
excel_db.close()

admin_id = [454589284, 1328571684, 1616544477]
users = []
swats = []
pvp = []
bg = []
# test_users = [1729215365, 454589284]

# Пользователи
for i in range(ex_user.max_row - 1):
    id_users = ex_user.cell(row=(i + 2), column=1).value
    if id_users is None:
        pass
    else:
        users.append(int(id_users))
# Test-chat
users.append(-1001700330817)

# Спецназ
for i in range(ex_swats.max_row - 1):
    id_swats = ex_swats.cell(row=(i + 2), column=1).value
    if id_swats is None:
        pass
    else:
        swats.append(int(id_swats))

# ПВП
for i in range(worksheet_build_pvp.max_row - 1):
    id_pvp = worksheet_build_pvp.cell(row=(i + 2), column=1).value
    pvp.append(id_pvp)

# БГ
for i in range(worksheet_build_bg.max_row - 1):
    id_bg = worksheet_build_bg.cell(row=(i + 2), column=1).value
    bg.append(id_bg)


def author(chat_id, us_list):
    strid = str(chat_id)
    for item in us_list:
        if str(item) == strid:
            return True
    return False
